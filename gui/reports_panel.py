"""
Panel de reportes. Horas trabajadas, retardos y exportacion a Excel.
Tema claro profesional - Paleta: Naranja #ea8511, Negro #1d120e, Gris #afaeb3
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime, date
import os

import database

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

DATA_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "data")

# --- Paleta NEVOX FARMA (tema claro) ---
NARANJA = "#ea8511"
NARANJA_LIGHT = "#fef7ed"
NARANJA_BORDER = "#fcd9a8"
NEGRO = "#1d120e"
TEXTO = "#1d120e"
TEXTO_SEC = "#6b6b70"
GRIS = "#afaeb3"
GRIS_BORDER = "#e0e0e2"
GRIS_BG = "#f5f5f6"
BLANCO = "#ffffff"
ROJO = "#dc2626"
AMARILLO = "#b45309"


class ReportsPanel(tk.Toplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title("Reportes - NEVOX FARMA")
        self.geometry("1000x680")
        self.minsize(850, 550)
        self.configure(bg=BLANCO)
        self.transient(parent)

        self._build_ui()

    def _build_ui(self):
        tk.Frame(self, bg=NARANJA, height=4).pack(fill=tk.X)

        header = tk.Frame(self, bg=BLANCO, padx=24, pady=16)
        header.pack(fill=tk.X)
        tk.Label(header, text="Reportes", bg=BLANCO, fg=NEGRO,
                  font=("Segoe UI", 18, "bold")).pack(side=tk.LEFT)

        tk.Frame(self, bg=GRIS_BORDER, height=1).pack(fill=tk.X)

        self.notebook = ttk.Notebook(self)
        self.notebook.pack(fill=tk.BOTH, expand=True, padx=24, pady=(12, 24))

        self.horas_frame = tk.Frame(self.notebook, bg=BLANCO)
        self.notebook.add(self.horas_frame, text="  Horas Trabajadas  ")
        self._build_horas_tab()

        self.retardos_frame = tk.Frame(self.notebook, bg=BLANCO)
        self.notebook.add(self.retardos_frame, text="  Retardos  ")
        self._build_retardos_tab()

        self.export_frame = tk.Frame(self.notebook, bg=BLANCO)
        self.notebook.add(self.export_frame, text="  Exportar a Excel  ")
        self._build_export_tab()

    def _make_date_toolbar(self, parent):
        toolbar = tk.Frame(parent, bg=GRIS_BG, padx=12, pady=10)
        toolbar.pack(fill=tk.X, pady=(0, 0))

        entry_kw = {"font": ("Segoe UI", 10), "bg": BLANCO, "fg": TEXTO,
                      "insertbackground": NARANJA, "relief": "flat",
                      "highlightthickness": 1, "highlightbackground": GRIS_BORDER,
                      "highlightcolor": NARANJA, "width": 12}

        tk.Label(toolbar, text="Desde:", bg=GRIS_BG, fg=TEXTO_SEC,
                  font=("Segoe UI", 10)).pack(side=tk.LEFT, padx=(0, 4))
        desde = tk.Entry(toolbar, **entry_kw)
        desde.insert(0, date.today().replace(day=1).isoformat())
        desde.pack(side=tk.LEFT, padx=(0, 14))

        tk.Label(toolbar, text="Hasta:", bg=GRIS_BG, fg=TEXTO_SEC,
                  font=("Segoe UI", 10)).pack(side=tk.LEFT, padx=(0, 4))
        hasta = tk.Entry(toolbar, **entry_kw)
        hasta.insert(0, date.today().isoformat())
        hasta.pack(side=tk.LEFT, padx=(0, 14))

        return toolbar, desde, hasta

    # --- Horas Trabajadas ---
    def _build_horas_tab(self):
        toolbar, self.horas_desde, self.horas_hasta = self._make_date_toolbar(self.horas_frame)
        ttk.Button(toolbar, text="Consultar", command=self._consultar_horas,
                    style="Accent.TButton").pack(side=tk.LEFT, padx=4)

        tree_frame = tk.Frame(self.horas_frame, bg=BLANCO)
        tree_frame.pack(fill=tk.BOTH, expand=True)

        columns = ("nombre", "departamento", "horas")
        self.horas_tree = ttk.Treeview(tree_frame, columns=columns, show="headings", selectmode="none")
        self.horas_tree.heading("nombre", text="EMPLEADO")
        self.horas_tree.heading("departamento", text="DEPARTAMENTO")
        self.horas_tree.heading("horas", text="HORAS TRABAJADAS")
        self.horas_tree.column("nombre", width=220)
        self.horas_tree.column("departamento", width=160)
        self.horas_tree.column("horas", width=140, anchor="center")

        scrollbar = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL, command=self.horas_tree.yview)
        self.horas_tree.configure(yscrollcommand=scrollbar.set)
        self.horas_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

    def _consultar_horas(self):
        desde = self.horas_desde.get().strip()
        hasta = self.horas_hasta.get().strip()

        try:
            date.fromisoformat(desde)
            date.fromisoformat(hasta)
        except ValueError:
            messagebox.showerror("Error", "Formato de fecha invalido. Usa AAAA-MM-DD.", parent=self)
            return

        for item in self.horas_tree.get_children():
            self.horas_tree.delete(item)

        empleados = database.listar_empleados()
        for emp in empleados:
            horas = database.calcular_horas_trabajadas(emp["id"], desde, hasta)
            self.horas_tree.insert("", tk.END, values=(emp["nombre"], emp["departamento"], f"{horas:.2f} hrs"))

    # --- Retardos ---
    def _build_retardos_tab(self):
        toolbar, self.ret_desde, self.ret_hasta = self._make_date_toolbar(self.retardos_frame)
        ttk.Button(toolbar, text="Consultar", command=self._consultar_retardos,
                    style="Accent.TButton").pack(side=tk.LEFT, padx=4)

        tree_frame = tk.Frame(self.retardos_frame, bg=BLANCO)
        tree_frame.pack(fill=tk.BOTH, expand=True)

        columns = ("nombre", "departamento", "fecha", "hora_prog", "hora_reg", "estado")
        self.ret_tree = ttk.Treeview(tree_frame, columns=columns, show="headings", selectmode="none")
        self.ret_tree.heading("nombre", text="EMPLEADO")
        self.ret_tree.heading("departamento", text="DEPARTAMENTO")
        self.ret_tree.heading("fecha", text="FECHA")
        self.ret_tree.heading("hora_prog", text="H. PROGRAMADA")
        self.ret_tree.heading("hora_reg", text="H. REGISTRO")
        self.ret_tree.heading("estado", text="ESTADO")
        self.ret_tree.column("nombre", width=170)
        self.ret_tree.column("departamento", width=130)
        self.ret_tree.column("fecha", width=100, anchor="center")
        self.ret_tree.column("hora_prog", width=100, anchor="center")
        self.ret_tree.column("hora_reg", width=100, anchor="center")
        self.ret_tree.column("estado", width=140, anchor="center")

        scrollbar = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL, command=self.ret_tree.yview)
        self.ret_tree.configure(yscrollcommand=scrollbar.set)
        self.ret_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

    def _consultar_retardos(self):
        desde = self.ret_desde.get().strip()
        hasta = self.ret_hasta.get().strip()

        try:
            date.fromisoformat(desde)
            date.fromisoformat(hasta)
        except ValueError:
            messagebox.showerror("Error", "Formato de fecha invalido. Usa AAAA-MM-DD.", parent=self)
            return

        for item in self.ret_tree.get_children():
            self.ret_tree.delete(item)

        retardos = database.obtener_retardos(desde, hasta)
        for r in retardos:
            estado = "Dentro de tolerancia" if r["con_tolerancia"] else "RETARDO"
            tag = "tolerancia" if r["con_tolerancia"] else "retardo"
            self.ret_tree.insert("", tk.END, values=(
                r["nombre"], r["departamento"], r["fecha"],
                r["hora_programada"], r["hora_registro"], estado
            ), tags=(tag,))

        self.ret_tree.tag_configure("retardo", foreground=ROJO)
        self.ret_tree.tag_configure("tolerancia", foreground=AMARILLO)

    # --- Exportar a Excel ---
    def _build_export_tab(self):
        frame = tk.Frame(self.export_frame, bg=BLANCO, padx=32, pady=32)
        frame.pack(fill=tk.BOTH, expand=True)

        tk.Label(frame, text="Exportar registros a archivo Excel",
                  bg=BLANCO, fg=NEGRO, font=("Segoe UI", 14, "bold")).grid(
            row=0, column=0, columnspan=3, sticky="w", pady=(0, 20))

        lbl_kw = {"bg": BLANCO, "fg": TEXTO_SEC, "font": ("Segoe UI", 11)}
        entry_kw = {"font": ("Segoe UI", 11), "bg": BLANCO, "fg": TEXTO,
                      "insertbackground": NARANJA, "relief": "flat",
                      "highlightthickness": 1, "highlightbackground": GRIS_BORDER,
                      "highlightcolor": NARANJA}

        tk.Label(frame, text="Desde:", **lbl_kw).grid(row=1, column=0, sticky="w", pady=10)
        self.exp_desde = tk.Entry(frame, width=14, **entry_kw)
        self.exp_desde.insert(0, date.today().replace(day=1).isoformat())
        self.exp_desde.grid(row=1, column=1, pady=10, padx=(12, 0), sticky="w")

        tk.Label(frame, text="Hasta:", **lbl_kw).grid(row=2, column=0, sticky="w", pady=10)
        self.exp_hasta = tk.Entry(frame, width=14, **entry_kw)
        self.exp_hasta.insert(0, date.today().isoformat())
        self.exp_hasta.grid(row=2, column=1, pady=10, padx=(12, 0), sticky="w")

        tk.Label(frame, text="Empleado:", **lbl_kw).grid(row=3, column=0, sticky="w", pady=10)

        empleados = database.listar_empleados()
        opciones = ["-- Todos --"] + [f"{e['id']} - {e['nombre']}" for e in empleados]
        self.var_emp_export = tk.StringVar(value=opciones[0])
        combo = ttk.Combobox(frame, textvariable=self.var_emp_export, values=opciones,
                              state="readonly", font=("Segoe UI", 11), width=28)
        combo.grid(row=3, column=1, pady=10, padx=(12, 0), sticky="w")

        ttk.Button(frame, text="Exportar a Excel", command=self._exportar_excel,
                    style="Accent.TButton").grid(row=4, column=0, columnspan=3, pady=28)

        if not HAS_OPENPYXL:
            warn_frame = tk.Frame(frame, bg="#fef2f2", padx=12, pady=8,
                                    highlightbackground="#fecaca", highlightthickness=1)
            warn_frame.grid(row=5, column=0, columnspan=3, sticky="ew", pady=8)
            tk.Label(warn_frame, text="openpyxl no esta instalado. Ejecuta: pip install openpyxl",
                      bg="#fef2f2", fg=ROJO, font=("Segoe UI", 10)).pack()

    def _exportar_excel(self):
        if not HAS_OPENPYXL:
            messagebox.showerror("Error", "openpyxl no esta instalado.\nEjecuta: pip install openpyxl", parent=self)
            return

        desde = self.exp_desde.get().strip()
        hasta = self.exp_hasta.get().strip()

        try:
            date.fromisoformat(desde)
            date.fromisoformat(hasta)
        except ValueError:
            messagebox.showerror("Error", "Formato de fecha invalido. Usa AAAA-MM-DD.", parent=self)
            return

        emp_sel = self.var_emp_export.get()
        emp_id = None
        if emp_sel != "-- Todos --":
            try:
                emp_id = int(emp_sel.split(" - ")[0])
            except (ValueError, IndexError):
                pass

        registros = database.obtener_registros_rango(desde, hasta, emp_id)

        if not registros:
            messagebox.showinfo("Info", "No hay registros en el rango seleccionado.", parent=self)
            return

        os.makedirs(DATA_DIR, exist_ok=True)
        default_name = f"registros_{desde}_{hasta}.xlsx"
        filepath = filedialog.asksaveasfilename(
            parent=self,
            initialdir=DATA_DIR,
            initialfile=default_name,
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
        )
        if not filepath:
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "Registros"

        # Estilos con colores de marca
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="ea8511", end_color="ea8511", fill_type="solid")
        title_font = Font(bold=True, size=14, color="1d120e")
        center = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin", color="e0e0e2"),
            right=Side(style="thin", color="e0e0e2"),
            top=Side(style="thin", color="e0e0e2"),
            bottom=Side(style="thin", color="e0e0e2"),
        )
        alt_fill = PatternFill(start_color="f7f7f8", end_color="f7f7f8", fill_type="solid")

        # Titulo
        ws.merge_cells("A1:E1")
        ws["A1"] = f"NEVOX FARMA - Registros del {desde} al {hasta}"
        ws["A1"].font = title_font
        ws["A1"].alignment = Alignment(vertical="center")
        ws.row_dimensions[1].height = 32

        # Encabezados
        headers = ["Fecha", "Hora", "Empleado", "Departamento", "Tipo"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = thin_border
        ws.row_dimensions[3].height = 28

        for i, reg in enumerate(registros, 4):
            dt = datetime.fromisoformat(reg["fecha_hora"])
            row_fill = alt_fill if i % 2 == 0 else None
            cells = [
                ws.cell(row=i, column=1, value=dt.strftime("%Y-%m-%d")),
                ws.cell(row=i, column=2, value=dt.strftime("%H:%M:%S")),
                ws.cell(row=i, column=3, value=reg["nombre"]),
                ws.cell(row=i, column=4, value=reg["departamento"]),
                ws.cell(row=i, column=5, value=reg["tipo"].upper()),
            ]
            for cell in cells:
                cell.border = thin_border
                if row_fill:
                    cell.fill = row_fill
            if reg["tipo"] == "entrada":
                cells[4].font = Font(color="16a34a", bold=True)
            else:
                cells[4].font = Font(color="ea8511", bold=True)

        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 28
        ws.column_dimensions["D"].width = 22
        ws.column_dimensions["E"].width = 12

        # Hoja resumen
        ws2 = wb.create_sheet("Horas Trabajadas")
        ws2.merge_cells("A1:C1")
        ws2["A1"] = f"NEVOX FARMA - Resumen de Horas ({desde} al {hasta})"
        ws2["A1"].font = title_font
        ws2["A1"].alignment = Alignment(vertical="center")
        ws2.row_dimensions[1].height = 32

        for col, h in enumerate(["Empleado", "Departamento", "Horas Trabajadas"], 1):
            cell = ws2.cell(row=3, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = thin_border
        ws2.row_dimensions[3].height = 28

        empleados = database.listar_empleados()
        for i, emp in enumerate(empleados, 4):
            horas = database.calcular_horas_trabajadas(emp["id"], desde, hasta)
            row_fill = alt_fill if i % 2 == 0 else None
            cells = [
                ws2.cell(row=i, column=1, value=emp["nombre"]),
                ws2.cell(row=i, column=2, value=emp["departamento"]),
                ws2.cell(row=i, column=3, value=f"{horas:.2f}"),
            ]
            for cell in cells:
                cell.border = thin_border
                if row_fill:
                    cell.fill = row_fill

        ws2.column_dimensions["A"].width = 28
        ws2.column_dimensions["B"].width = 22
        ws2.column_dimensions["C"].width = 20

        wb.save(filepath)
        messagebox.showinfo("Exito", f"Archivo exportado correctamente:\n{filepath}", parent=self)
