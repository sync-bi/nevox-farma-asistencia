"""
NEVOX FARMA - Sistema de Control de Asistencia
Aplicacion Flask para Vercel con Supabase (REST API directo).
Archivo unico: database + QR + rutas.
"""

import os
import hashlib
import hmac
import json
import secrets
import time
import io
import base64
import traceback
from io import BytesIO
from functools import wraps
from datetime import datetime, date, timedelta, timezone

import requests as _http
from flask import (
    Flask, request, render_template, jsonify, session,
    redirect, url_for, send_file,
)
import qrcode
from PIL import Image

# ============================================================
# CONFIG
# ============================================================

SUPABASE_URL = os.environ.get("SUPABASE_URL", "")
SUPABASE_KEY = os.environ.get("SUPABASE_KEY", "")
BASE_URL = os.environ.get("BASE_URL", "http://localhost:5000").rstrip("/")
QR_ROTATION_INTERVAL = 30

# ------------------------------------------------------------
# ZONA HORARIA
# Supabase guarda fecha_hora en UTC. El servidor (Vercel) tambien corre en
# UTC, por eso datetime.now()/date.today() dan la hora equivocada. Todo lo
# que se muestre o filtre por "hoy" debe pasar por estos helpers (UTC-5).
# ------------------------------------------------------------
LOCAL_TZ = timezone(timedelta(hours=-5))  # Lima / Bogota / Quito


def now_local():
    return datetime.now(timezone.utc).astimezone(LOCAL_TZ)


def today_local():
    return now_local().date()


def to_local(s):
    """Convierte un fecha_hora almacenado (UTC) a datetime local con tzinfo."""
    dt = datetime.fromisoformat(s)
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(LOCAL_TZ)


def local_day_bounds_utc(desde, hasta=None):
    """Dado un rango de fechas LOCALES (YYYY-MM-DD) devuelve los limites
    equivalentes en UTC para filtrar la columna fecha_hora en Supabase."""
    hasta = hasta or desde
    ini = datetime.fromisoformat(f"{desde}T00:00:00").replace(tzinfo=LOCAL_TZ)
    fin = datetime.fromisoformat(f"{hasta}T23:59:59").replace(tzinfo=LOCAL_TZ)
    return (
        ini.astimezone(timezone.utc).strftime("%Y-%m-%dT%H:%M:%S"),
        fin.astimezone(timezone.utc).strftime("%Y-%m-%dT%H:%M:%S"),
    )


def _sb_headers(prefer=None):
    h = {
        "apikey": SUPABASE_KEY,
        "Authorization": f"Bearer {SUPABASE_KEY}",
        "Content-Type": "application/json",
    }
    if prefer:
        h["Prefer"] = prefer
    return h


def _sb_get(table, select="*", filters=None, order=None, limit=None):
    params = [("select", select)]
    if filters:
        params.extend(filters)
    if order:
        params.append(("order", order))
    if limit:
        params.append(("limit", str(limit)))
    r = _http.get(f"{SUPABASE_URL}/rest/v1/{table}", params=params, headers=_sb_headers())
    r.raise_for_status()
    return r.json()


def _sb_post(table, data, prefer="return=representation"):
    r = _http.post(f"{SUPABASE_URL}/rest/v1/{table}", json=data, headers=_sb_headers(prefer))
    r.raise_for_status()
    return r.json() if prefer and "return" in prefer else None


def _sb_upsert(table, data):
    r = _http.post(
        f"{SUPABASE_URL}/rest/v1/{table}", json=data,
        headers=_sb_headers("resolution=merge-duplicates,return=representation"),
    )
    r.raise_for_status()
    return r.json()


def _sb_patch(table, data, filters):
    r = _http.patch(
        f"{SUPABASE_URL}/rest/v1/{table}", json=data,
        params=filters, headers=_sb_headers("return=representation"),
    )
    r.raise_for_status()
    return r.json()


def _sb_delete(table, filters):
    r = _http.delete(
        f"{SUPABASE_URL}/rest/v1/{table}",
        params=filters, headers=_sb_headers(),
    )
    r.raise_for_status()


def _sb_rpc(fn_name, data):
    r = _http.post(
        f"{SUPABASE_URL}/rest/v1/rpc/{fn_name}", json=data,
        headers=_sb_headers(),
    )
    r.raise_for_status()
    return r.json()


# ============================================================
# DATABASE FUNCTIONS
# ============================================================

def db_get_config(clave):
    data = _sb_get("configuracion", select="valor", filters=[("clave", f"eq.{clave}")])
    return data[0]["valor"] if data else None


def db_set_config(clave, valor):
    _sb_upsert("configuracion", {"clave": clave, "valor": valor})


def db_verificar_password(password):
    hashed = hashlib.sha256(password.encode()).hexdigest()
    return hashed == db_get_config("admin_password")


def db_cambiar_password(nuevo):
    hashed = hashlib.sha256(nuevo.encode()).hexdigest()
    db_set_config("admin_password", hashed)


def db_crear_empleado(nombre, departamento="", hora_entrada="09:00", hora_salida="18:00"):
    result = _sb_post("empleados", {
        "nombre": nombre, "departamento": departamento,
        "hora_entrada": hora_entrada, "hora_salida": hora_salida,
    })
    return result[0]["id"]


def _fix_activo(row):
    row["activo"] = 1 if row.get("activo") else 0
    return row


def db_obtener_empleado(empleado_id):
    data = _sb_get("empleados", filters=[("id", f"eq.{empleado_id}")])
    return _fix_activo(data[0]) if data else None


def db_obtener_empleado_por_token(token):
    data = _sb_get("empleados", filters=[
        ("token_dispositivo", f"eq.{token}"), ("activo", "eq.true"),
    ])
    return _fix_activo(data[0]) if data else None


def db_listar_empleados(solo_activos=True):
    filters = [("activo", "eq.true")] if solo_activos else []
    data = _sb_get("empleados", filters=filters, order="nombre.asc")
    return [_fix_activo(r) for r in data]


def db_actualizar_empleado(emp_id, **kwargs):
    campos = {}
    for k in ["nombre", "departamento", "hora_entrada", "hora_salida"]:
        if k in kwargs and kwargs[k] is not None:
            campos[k] = kwargs[k]
    if "activo" in kwargs and kwargs["activo"] is not None:
        campos["activo"] = bool(kwargs["activo"])
    if campos:
        _sb_patch("empleados", campos, [("id", f"eq.{emp_id}")])


def db_vincular(emp_id, token):
    _sb_patch("empleados", {"token_dispositivo": token}, [("id", f"eq.{emp_id}")])


def db_desvincular(emp_id):
    _sb_patch("empleados", {"token_dispositivo": None}, [("id", f"eq.{emp_id}")])


def db_registrar_asistencia(emp_id, tipo, token_usado=None):
    filas = _sb_post("registros", {"empleado_id": emp_id, "tipo": tipo, "token_usado": token_usado})
    return filas[0] if filas else None


def db_ultimo_registro(emp_id, fecha=None):
    fecha = fecha or today_local().isoformat()
    ini, fin = local_day_bounds_utc(fecha)
    data = _sb_get("registros", filters=[
        ("empleado_id", f"eq.{emp_id}"),
        ("fecha_hora", f"gte.{ini}"),
        ("fecha_hora", f"lte.{fin}"),
    ], order="fecha_hora.desc", limit=1)
    return data[0] if data else None


def _minutos_del_dia(hhmm):
    h, m = map(int, hhmm.split(":")[:2])
    return h * 60 + m


def _punto_medio(hhmm_a, hhmm_b):
    a = _minutos_del_dia(hhmm_a)
    b = _minutos_del_dia(hhmm_b)
    med = (a + b) // 2
    return f"{med // 60:02d}:{med % 60:02d}"


def db_registros_hoy_empleado(emp_id, fecha=None):
    """Marcas del empleado en el dia local indicado, en orden ascendente."""
    fecha = fecha or today_local().isoformat()
    ini, fin = local_day_bounds_utc(fecha)
    return _sb_get("registros", filters=[
        ("empleado_id", f"eq.{emp_id}"),
        ("fecha_hora", f"gte.{ini}"),
        ("fecha_hora", f"lte.{fin}"),
    ], order="fecha_hora.asc")


def tipo_por_hora(momento, horario=None):
    """Tipo de la PRIMERA marca del dia, decidido por la hora: antes del punto
    medio de la jornada es entrada, despues es salida."""
    turno = (horario or db_get_horario_semanal())[str(momento.date().weekday())]
    if not turno:
        return "entrada"  # dia no laboral: no hay jornada de referencia
    return "entrada" if momento.strftime("%H:%M") < _punto_medio(turno["entrada"], turno["salida"]) else "salida"


def db_siguiente_tipo(emp_id, momento=None):
    """Tipo del proximo registro.

    Si el empleado ya marco hoy se alterna a partir del ultimo registro. Si es
    la PRIMERA marca del dia no hay nada con que alternar, asi que se decide
    por la hora: antes del punto medio de la jornada es entrada, despues es
    salida. Sin esto, a quien se le olvidaba marcar la entrada en la manana se
    le registraba la salida de la tarde como "entrada", lo que ademas de
    perder el dia generaba un retardo falso de varias horas.
    """
    ultimo = db_ultimo_registro(emp_id)
    if ultimo is not None:
        return "entrada" if ultimo["tipo"] == "salida" else "salida"
    return tipo_por_hora(momento or now_local())


# Ventana anti-rebote: /checkin dispara el registro en cada carga de la pagina,
# asi que una recarga, el boton atras, un doble escaneo (el QR vale hasta 60s) o
# la precarga que hacen algunas apps de camara metian un segundo registro que,
# al alternar entrada/salida, marcaba la salida de inmediato. Dentro de esta
# ventana se devuelve el registro anterior en vez de crear uno nuevo.
ANTIRREBOTE_DEFAULT = 90  # segundos


def db_get_antirrebote():
    try:
        v = int(db_get_config("checkin_antirrebote_segundos"))
        return v if v >= 0 else ANTIRREBOTE_DEFAULT
    except (TypeError, ValueError):
        return ANTIRREBOTE_DEFAULT




def _flatten_registros(data):
    registros = []
    for r in data:
        emp = r.pop("empleados", {}) or {}
        r["nombre"] = emp.get("nombre", "")
        r["departamento"] = emp.get("departamento", "")
        if r.get("fecha_hora"):
            loc = to_local(r["fecha_hora"])
            r["fecha_hora"] = loc.isoformat()        # ya en hora local (-05:00)
            r["hora"] = loc.strftime("%H:%M:%S")     # listo para mostrar
            r["fecha"] = loc.strftime("%d/%m/%Y")
        registros.append(r)
    return registros


def db_registros_dia(fecha=None):
    fecha = fecha or today_local().isoformat()
    ini, fin = local_day_bounds_utc(fecha)
    data = _sb_get("registros", select="*,empleados(nombre,departamento)", filters=[
        ("fecha_hora", f"gte.{ini}"),
        ("fecha_hora", f"lte.{fin}"),
    ], order="fecha_hora.desc")
    return _flatten_registros(data)


def db_registros_rango(desde, hasta, emp_id=None):
    ini, fin = local_day_bounds_utc(desde, hasta)
    filters = [
        ("fecha_hora", f"gte.{ini}"),
        ("fecha_hora", f"lte.{fin}"),
    ]
    if emp_id:
        filters.append(("empleado_id", f"eq.{emp_id}"))
    data = _sb_get("registros", select="*,empleados(nombre,departamento)", filters=filters, order="fecha_hora.asc")
    return _flatten_registros(data)


# ------------------------------------------------------------
# HORARIO SEMANAL Y HORAS EXTRAS
#
# El horario vive en la tabla configuracion (clave "horario_semanal") como
# JSON, asi no hace falta migrar el esquema de Supabase:
#   {"0": {"entrada": "07:00", "salida": "16:00"}, ..., "6": null}
# donde 0 = lunes ... 6 = domingo y null = dia no laboral.
#
# Hora extra = tiempo efectivamente trabajado DESPUES de la hora de salida
# programada de ese dia. Se calcula sobre los pares entrada/salida reales,
# de modo que salir a almorzar y volver no infla el resultado. Llegar antes
# de la hora de entrada no genera extra. En un dia no laboral (sabado /
# domingo) cuenta todo el tiempo trabajado, pero marcado aparte.
# ------------------------------------------------------------

DIAS_SEMANA = ["Lunes", "Martes", "Miercoles", "Jueves", "Viernes", "Sabado", "Domingo"]

SIN_AREA = "Sin area"  # etiqueta para empleados sin departamento asignado

# En NEVOX FARMA el almuerzo NO se marca: cada dia son exactamente dos marcas,
# la entrada y la salida. Cualquier marca adicional se rechaza en el check-in y
# los dias historicos que tengan mas quedan senalados para revision.
MARCAS_ESPERADAS = 2

# Una jornada por debajo de esto no es un dia real: casi siempre es el rastro
# de un doble escaneo (entrada y salida con minutos de diferencia). Se marca
# para revision, no se descarta.
JORNADA_MINIMA_DEFAULT = 30  # minutos

# Por debajo de esto la pareja entrada/salida es directamente un duplicado: no
# existe una jornada real de menos de 5 minutos. El check-in lo usa para dejar
# que la persona marque su salida de verdad en vez de bloquearla.
DUPLICADO_MAX_MINUTOS = 5

# Horario "Apoyo" informado por NEVOX FARMA (28/07/2026).
HORARIO_SEMANAL_DEFAULT = {
    "0": {"entrada": "07:00", "salida": "16:00"},   # lunes
    "1": {"entrada": "07:00", "salida": "16:00"},   # martes
    "2": {"entrada": "07:00", "salida": "16:30"},   # miercoles
    "3": {"entrada": "07:00", "salida": "16:00"},   # jueves
    "4": {"entrada": "07:00", "salida": "16:00"},   # viernes
    "5": None,                                      # sabado
    "6": None,                                      # domingo
}

EXTRAS_MINIMO_DEFAULT = 15    # minutos: por debajo de esto no se cuenta extra
EXTRAS_REDONDEO_DEFAULT = 15  # minutos: bloque de redondeo hacia abajo


def _hhmm_valido(v):
    try:
        h, m = str(v).split(":")[:2]
        return 0 <= int(h) <= 23 and 0 <= int(m) <= 59
    except (ValueError, AttributeError):
        return False


def normalizar_horario(raw):
    """Valida y completa un horario semanal. Devuelve claves '0'..'6', donde
    cada valor es {'entrada','salida'} o None si el dia no es laboral."""
    horario = {}
    for i in range(7):
        v = (raw or {}).get(str(i))
        if isinstance(v, dict) and _hhmm_valido(v.get("entrada")) and _hhmm_valido(v.get("salida")):
            horario[str(i)] = {"entrada": v["entrada"][:5], "salida": v["salida"][:5]}
        else:
            horario[str(i)] = None
    return horario


def db_get_horario_semanal():
    raw = db_get_config("horario_semanal")
    if not raw:
        return normalizar_horario(HORARIO_SEMANAL_DEFAULT)
    try:
        return normalizar_horario(json.loads(raw))
    except (ValueError, TypeError):
        return normalizar_horario(HORARIO_SEMANAL_DEFAULT)


def db_set_horario_semanal(horario):
    db_set_config("horario_semanal", json.dumps(normalizar_horario(horario)))


def db_get_reglas_extras():
    def _int(clave, default):
        try:
            v = int(db_get_config(clave))
            return v if v >= 0 else default
        except (TypeError, ValueError):
            return default
    return {
        "minimo": _int("extras_minimo_minutos", EXTRAS_MINIMO_DEFAULT),
        "redondeo": _int("extras_redondeo_minutos", EXTRAS_REDONDEO_DEFAULT),
    }


def db_get_jornada_minima():
    try:
        v = int(db_get_config("jornada_minima_minutos"))
        return v if v >= 0 else JORNADA_MINIMA_DEFAULT
    except (TypeError, ValueError):
        return JORNADA_MINIMA_DEFAULT


def _jornadas_por_dia(registros, jornada_minima=None):
    """Agrupa los registros (ya en hora local, orden ascendente) en pares
    entrada/salida por empleado y dia."""
    if jornada_minima is None:
        jornada_minima = db_get_jornada_minima()
    dias = {}
    for r in registros:
        dt = datetime.fromisoformat(r["fecha_hora"])
        key = (r["empleado_id"], dt.strftime("%Y-%m-%d"))
        d = dias.setdefault(key, {
            "nombre": r.get("nombre", ""),
            "departamento": r.get("departamento", ""),
            "pares": [], "pendiente": None, "marcas": 0,
            "primera_entrada": None, "ultima_salida": None,
            "entradas_sin_salida": 0, "salidas_sin_entrada": 0,
        })
        d["marcas"] += 1
        if r["tipo"] == "entrada":
            if d["pendiente"] is not None:
                # Dos entradas seguidas: a la anterior le falto su salida.
                d["entradas_sin_salida"] += 1
            d["pendiente"] = dt
            if d["primera_entrada"] is None:
                d["primera_entrada"] = dt
        elif r["tipo"] == "salida":
            if d["pendiente"] is not None:
                d["pares"].append((d["pendiente"], dt))
                d["pendiente"] = None
            else:
                # Salida sin entrada previa: antes se descartaba en silencio.
                d["salidas_sin_entrada"] += 1
            d["ultima_salida"] = dt

    for d in dias.values():
        # La entrada que quedo sin cerrar; sirve para avisar en el dashboard.
        d["abierta_desde"] = d.pop("pendiente")
        if d["abierta_desde"] is not None:
            d["entradas_sin_salida"] += 1
        d["incompleto"] = bool(d["entradas_sin_salida"] or d["salidas_sin_entrada"])
        # En NEVOX no se marca el almuerzo: se esperan exactamente 2 marcas al
        # dia. Mas de dos es una anomalia que hay que revisar aunque el dia
        # cierre bien (tipico de los escaneos duplicados que ya no ocurren).
        d["exceso_marcas"] = max(0, d["marcas"] - MARCAS_ESPERADAS)
        d["trabajado_min"] = sum((f - i).total_seconds() for i, f in d["pares"]) / 60
        # Dia que empareja bien pero dura casi nada: pasaba como limpio y en
        # realidad es un doble escaneo que se comio la jornada entera.
        d["jornada_corta"] = bool(d["pares"]) and d["trabajado_min"] < jornada_minima

        faltantes = []
        if d["entradas_sin_salida"]:
            faltantes.append("Falta salida")
        if d["salidas_sin_entrada"]:
            faltantes.append("Falta entrada")
        if d["exceso_marcas"]:
            faltantes.append(f"{d['marcas']} marcas (se esperan {MARCAS_ESPERADAS})")
        if d["jornada_corta"]:
            faltantes.append(f"Jornada de {int(round(d['trabajado_min']))} min (revisar)")
        d["motivo"] = " y ".join(faltantes)
        d["revisar"] = bool(d["incompleto"] or d["exceso_marcas"] or d["jornada_corta"])
    return dias


def _minutos_extra(pares, salida_prog):
    """Minutos trabajados despues de salida_prog. Si salida_prog es None
    (dia no laboral) cuenta todo el tiempo trabajado."""
    total = 0.0
    for ini, fin in pares:
        desde = ini if salida_prog is None else max(ini, salida_prog)
        seg = (fin - desde).total_seconds()
        if seg > 0:
            total += seg
    return total / 60


def _aplicar_reglas_extras(minutos, reglas):
    """Aplica umbral minimo y redondeo hacia abajo. Devuelve minutos enteros."""
    if minutos < reglas["minimo"]:
        return 0
    if reglas["redondeo"] > 0:
        return (int(minutos) // reglas["redondeo"]) * reglas["redondeo"]
    return int(minutos)


def _resumen_vacio(emp_id, nombre, departamento):
    return {
        "empleado_id": emp_id, "nombre": nombre, "departamento": departamento,
        "horas": 0.0, "extras_horas": 0.0, "extras_habil_horas": 0.0,
        "extras_no_laboral_horas": 0.0, "dias_con_extra": 0, "dias_incompletos": 0,
    }


def db_resumen_periodo(desde, hasta, emp_id=None, departamento=None):
    """Detalle diario (horas trabajadas y extras) y resumen por empleado
    para el rango indicado. Una sola consulta a Supabase para todo."""
    horario = db_get_horario_semanal()
    reglas = db_get_reglas_extras()
    dias = _jornadas_por_dia(db_registros_rango(desde, hasta, emp_id))
    if departamento:
        dias = {k: v for k, v in dias.items() if (v["departamento"] or SIN_AREA) == departamento}

    detalle = []
    for (eid, fecha), d in sorted(dias.items(), key=lambda kv: ((kv[1]["departamento"] or SIN_AREA), kv[1]["nombre"], kv[0][1])):
        fecha_d = date.fromisoformat(fecha)
        wd = fecha_d.weekday()
        turno = horario[str(wd)]
        salida_prog = None
        if turno:
            h, m = map(int, turno["salida"].split(":"))
            salida_prog = datetime.combine(
                fecha_d, datetime.min.time().replace(hour=h, minute=m), tzinfo=LOCAL_TZ
            )

        trabajado = d["trabajado_min"] / 60
        bruto = _minutos_extra(d["pares"], salida_prog)
        extra = _aplicar_reglas_extras(bruto, reglas)
        detalle.append({
            "empleado_id": eid,
            "nombre": d["nombre"],
            "departamento": d["departamento"] or SIN_AREA,
            "fecha": fecha,
            "fecha_fmt": fecha_d.strftime("%d/%m/%Y"),
            "dia": DIAS_SEMANA[wd],
            "entrada": d["primera_entrada"].strftime("%H:%M") if d["primera_entrada"] else "",
            "salida": d["ultima_salida"].strftime("%H:%M") if d["ultima_salida"] else "",
            "entrada_programada": turno["entrada"] if turno else "",
            "salida_programada": turno["salida"] if turno else "",
            "horas_trabajadas": round(trabajado, 2),
            "extra_bruto_min": int(round(bruto)),
            "extra_min": extra,
            "extra_horas": round(extra / 60, 2),
            "no_laboral": turno is None,
            "incompleto": d["incompleto"],
            "revisar": d["revisar"],
            "motivo": d["motivo"],
            "marcas": d["marcas"],
        })

    resumen = {}
    for e in db_listar_empleados():
        if emp_id and e["id"] != emp_id:
            continue
        area = e["departamento"] or SIN_AREA
        if departamento and area != departamento:
            continue
        resumen[e["id"]] = _resumen_vacio(e["id"], e["nombre"], area)
    for d in detalle:
        # Un empleado desactivado despues de trabajar igual debe aparecer.
        r = resumen.setdefault(
            d["empleado_id"], _resumen_vacio(d["empleado_id"], d["nombre"], d["departamento"])
        )
        r["horas"] += d["horas_trabajadas"]
        r["extras_horas"] += d["extra_horas"]
        if d["no_laboral"]:
            r["extras_no_laboral_horas"] += d["extra_horas"]
        else:
            r["extras_habil_horas"] += d["extra_horas"]
        if d["extra_min"] > 0:
            r["dias_con_extra"] += 1
        if d["incompleto"]:
            r["dias_incompletos"] += 1

    for r in resumen.values():
        for k in ["horas", "extras_horas", "extras_habil_horas", "extras_no_laboral_horas"]:
            r[k] = round(r[k], 2)

    return {
        "detalle": detalle,
        "resumen": sorted(resumen.values(), key=lambda r: (r["departamento"], r["nombre"])),
        "horario": horario,
        "reglas": reglas,
    }


def db_listar_areas():
    areas = {(e["departamento"] or SIN_AREA) for e in db_listar_empleados()}
    return sorted(areas)


def _minutos_entre(hhmm_ini, hhmm_fin):
    return _minutos_del_dia(hhmm_fin) - _minutos_del_dia(hhmm_ini)


def db_retardos(desde, hasta, departamento=None):
    """Retardos del periodo. La hora de entrada sale del horario semanal
    (misma fuente que las horas extras), no de empleados.hora_entrada.

    Devuelve (retardos, sin_corregir): el segundo es el numero de dias que se
    dejaron fuera por tener la marca mal tipada. Ver el filtro del punto medio
    mas abajo.
    """
    emp_map = {e["id"]: e for e in db_listar_empleados()}
    horario = db_get_horario_semanal()
    tol = int(db_get_config("tolerancia_minutos") or "15")

    # Primera ENTRADA (hora local) por empleado y dia, calculada a partir de
    # los registros ya convertidos a UTC-5. Asi la comparacion con la hora
    # programada (que es local) es correcta.
    primeras = {}  # (empleado_id, fecha_local) -> "HH:MM"
    for r in db_registros_rango(desde, hasta):
        if r["tipo"] != "entrada":
            continue
        loc = datetime.fromisoformat(r["fecha_hora"])  # ya en hora local
        key = (r["empleado_id"], loc.strftime("%Y-%m-%d"))
        hhmm = loc.strftime("%H:%M")
        if key not in primeras or hhmm < primeras[key]:
            primeras[key] = hhmm

    retardos = []
    sin_corregir = 0
    for (emp_id, fecha), hora_reg in sorted(primeras.items()):
        emp = emp_map.get(emp_id)
        if not emp:
            continue
        if departamento and emp["departamento"] != departamento:
            continue
        fecha_d = date.fromisoformat(fecha)
        turno = horario[str(fecha_d.weekday())]
        if not turno:
            continue  # dia no laboral: trabajar ahi no es un retardo
        hora_limite = turno["entrada"]
        if hora_reg <= hora_limite:
            continue
        # Una "entrada" despues del punto medio de la jornada no es una llegada
        # tarde de 10 horas: es la salida de quien olvido marcar en la manana,
        # guardada como entrada por la alternancia vieja. Se usa el mismo punto
        # medio con que el check-in decide el tipo de la primera marca, asi el
        # reporte no contradice a la app. Esos dias no se pierden: salen en
        # Admin -> Corregir Registros con su motivo, y aqui se cuentan aparte.
        if hora_reg >= _punto_medio(turno["entrada"], turno["salida"]):
            sin_corregir += 1
            continue
        h, m = map(int, hora_limite.split(":"))
        lim = (
            datetime.combine(fecha_d, datetime.min.time().replace(hour=h, minute=m))
            + timedelta(minutes=tol)
        ).strftime("%H:%M")
        retardos.append({
            "empleado_id": emp_id, "nombre": emp["nombre"],
            "departamento": emp["departamento"] or "Sin area",
            "fecha": fecha, "fecha_fmt": fecha_d.strftime("%d/%m/%Y"),
            "dia": DIAS_SEMANA[fecha_d.weekday()],
            "hora_programada": hora_limite, "hora_registro": hora_reg,
            "minutos_tarde": _minutos_entre(hora_limite, hora_reg),
            "con_tolerancia": hora_reg <= lim,
        })
    return retardos, sin_corregir


# ------------------------------------------------------------
# CORRECCION MANUAL DE REGISTROS
# Un olvido al marcar deja el dia sin cerrar y contamina horas y retardos.
# Estas funciones permiten al admin agregar la marca que falta o arreglar
# una mal clasificada, sin tocar la base a mano.
# ------------------------------------------------------------

def _fecha_hora_utc(fecha, hora):
    """(YYYY-MM-DD, HH:MM) local -> timestamp UTC para guardar en Supabase."""
    hhmmss = hora if len(hora) > 5 else f"{hora}:00"
    dt = datetime.fromisoformat(f"{fecha}T{hhmmss}").replace(tzinfo=LOCAL_TZ)
    return dt.astimezone(timezone.utc).strftime("%Y-%m-%dT%H:%M:%S+00:00")


def db_crear_registro(emp_id, fecha, hora, tipo):
    filas = _sb_post("registros", {
        "empleado_id": emp_id, "tipo": tipo,
        "fecha_hora": _fecha_hora_utc(fecha, hora),
        "token_usado": "correccion-manual",
    })
    return filas[0] if filas else None


def db_actualizar_registro(reg_id, fecha=None, hora=None, tipo=None):
    campos = {}
    if tipo:
        campos["tipo"] = tipo
    if fecha and hora:
        campos["fecha_hora"] = _fecha_hora_utc(fecha, hora)
    if campos:
        _sb_patch("registros", campos, [("id", f"eq.{reg_id}")])


def db_eliminar_registro(reg_id):
    _sb_delete("registros", [("id", f"eq.{reg_id}")])


def db_mover_salida(reg_id, momento):
    """Reubica una salida duplicada a la hora real de salida. Queda marcada en
    token_usado para que se vea de donde salio."""
    _sb_patch("registros", {
        "fecha_hora": momento.astimezone(timezone.utc).strftime("%Y-%m-%dT%H:%M:%S+00:00"),
        "token_usado": "auto-correccion-duplicado",
    }, [("id", f"eq.{reg_id}")])


def db_sin_cerrar(fecha=None):
    """Quien tiene una entrada abierta hoy. Mientras no pase la hora de salida
    programada es normal (estan trabajando); despues, es una salida sin marcar."""
    fecha = fecha or today_local().isoformat()
    fecha_d = date.fromisoformat(fecha)
    turno = db_get_horario_semanal()[str(fecha_d.weekday())]
    salida_prog = turno["salida"] if turno else None
    ahora = now_local()
    vencido = bool(salida_prog and ahora.strftime("%H:%M") > salida_prog)

    pendientes = []
    for (eid, _f), d in _jornadas_por_dia(db_registros_rango(fecha, fecha)).items():
        if d["abierta_desde"] is None:
            continue
        pendientes.append({
            "empleado_id": eid, "nombre": d["nombre"],
            "departamento": d["departamento"] or SIN_AREA,
            "desde": d["abierta_desde"].strftime("%H:%M"),
            "minutos": max(0, int((ahora - d["abierta_desde"]).total_seconds() // 60)),
            "vencido": vencido,
        })
    pendientes.sort(key=lambda p: p["nombre"])
    return {
        "pendientes": pendientes,
        "total": len(pendientes),
        "vencidos": len(pendientes) if vencido else 0,
        "salida_programada": salida_prog or "",
        "fecha": fecha_d.strftime("%d/%m/%Y"),
    }


def db_dias_por_corregir(desde, hasta, emp_id=None):
    """Dias con marcas faltantes, listos para que el admin los arregle."""
    dias = _jornadas_por_dia(db_registros_rango(desde, hasta, emp_id))
    pendientes = []
    for (eid, fecha), d in dias.items():
        if not d["revisar"]:
            continue
        fecha_d = date.fromisoformat(fecha)
        pendientes.append({
            "empleado_id": eid, "nombre": d["nombre"],
            "departamento": d["departamento"] or SIN_AREA,
            "fecha": fecha, "fecha_fmt": fecha_d.strftime("%d/%m/%Y"),
            "dia": DIAS_SEMANA[fecha_d.weekday()],
            "motivo": d["motivo"], "marcas": d["marcas"],
            "primera_entrada": d["primera_entrada"].strftime("%H:%M") if d["primera_entrada"] else "",
            "ultima_salida": d["ultima_salida"].strftime("%H:%M") if d["ultima_salida"] else "",
        })
    return sorted(pendientes, key=lambda p: (p["departamento"], p["nombre"], p["fecha"]))


def db_limpiar_registros():
    _sb_delete("registros", [("id", "neq.0")])


def db_limpiar_todo():
    _sb_delete("registros", [("id", "neq.0")])
    _sb_delete("empleados", [("id", "neq.0")])


# ============================================================
# QR / TOKEN FUNCTIONS
# ============================================================

def _secret():
    return db_get_config("secret_key")


def qr_token():
    secret = _secret()
    slot = int(time.time()) // QR_ROTATION_INTERVAL
    firma = hmac.new(secret.encode(), f"qr:{slot}".encode(), hashlib.sha256).hexdigest()
    return f"{slot}:{firma}"


def qr_validar(token):
    try:
        slot_str, firma = token.split(":")
        slot_r = int(slot_str)
    except (ValueError, AttributeError):
        return False
    secret = _secret()
    slot_now = int(time.time()) // QR_ROTATION_INTERVAL
    for s in [slot_now, slot_now - 1]:
        expected = hmac.new(secret.encode(), f"qr:{s}".encode(), hashlib.sha256).hexdigest()
        if s == slot_r and hmac.compare_digest(firma, expected):
            return True
    return False


def device_token(emp_id):
    secret = _secret()
    rand = secrets.token_hex(16)
    firma = hmac.new(secret.encode(), f"device:{emp_id}:{rand}".encode(), hashlib.sha256).hexdigest()
    return f"dev:{emp_id}:{rand}:{firma}"


def device_validar(token):
    try:
        prefix, emp_id, rand, firma = token.split(":")
        if prefix != "dev":
            return None
    except (ValueError, AttributeError):
        return None
    secret = _secret()
    expected = hmac.new(secret.encode(), f"device:{emp_id}:{rand}".encode(), hashlib.sha256).hexdigest()
    return int(emp_id) if hmac.compare_digest(firma, expected) else None


def reg_token(emp_id):
    secret = _secret()
    rand = secrets.token_hex(16)
    firma = hmac.new(secret.encode(), f"reg:{emp_id}:{rand}".encode(), hashlib.sha256).hexdigest()
    return f"reg:{emp_id}:{rand}:{firma}"


def reg_validar(token):
    try:
        prefix, emp_id, rand, firma = token.split(":")
        if prefix != "reg":
            return None
    except (ValueError, AttributeError):
        return None
    secret = _secret()
    expected = hmac.new(secret.encode(), f"reg:{emp_id}:{rand}".encode(), hashlib.sha256).hexdigest()
    return int(emp_id) if hmac.compare_digest(firma, expected) else None


def qr_base64(data, size=8):
    qr = qrcode.QRCode(version=1, error_correction=qrcode.constants.ERROR_CORRECT_M, box_size=size, border=2)
    qr.add_data(data)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white").convert("RGB")
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    buf.seek(0)
    return base64.b64encode(buf.read()).decode("utf-8")


def qr_checkin_url():
    return f"{BASE_URL}/checkin?token={qr_token()}"


def qr_registro_url(emp_id):
    return f"{BASE_URL}/registro-dispositivo?token={reg_token(emp_id)}"


# ============================================================
# FLASK APP
# ============================================================

TEMPLATE_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "templates")

app = Flask(__name__, template_folder=TEMPLATE_DIR)
app.secret_key = os.environ.get("SECRET_KEY", "dev-fallback-change-me")


@app.errorhandler(Exception)
def handle_error(e):
    return jsonify({"error": str(e), "type": type(e).__name__, "trace": traceback.format_exc()}), 500


@app.route("/api/health")
def health():
    return jsonify({"status": "ok", "time": now_local().isoformat()})


def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("admin"):
            if request.is_json or request.path.startswith("/api/"):
                return jsonify({"ok": False, "mensaje": "No autorizado."}), 401
            return redirect(url_for("admin_login"))
        return f(*args, **kwargs)
    return decorated


# --- DASHBOARD ---
@app.route("/")
def index():
    return render_template("dashboard.html")


@app.route("/api/qr")
def api_qr():
    url = qr_checkin_url()
    b64 = qr_base64(url)
    rem = QR_ROTATION_INTERVAL - (int(time.time()) % QR_ROTATION_INTERVAL)
    return jsonify({"qr_base64": b64, "remaining_seconds": rem, "timestamp": now_local().strftime("%H:%M:%S")})


@app.route("/api/pendientes-hoy")
def api_pendientes_hoy():
    return jsonify(db_sin_cerrar())


@app.route("/api/registros-hoy")
def api_registros_hoy():
    regs = db_registros_dia()
    ent = sum(1 for r in regs if r["tipo"] == "entrada")
    sal = sum(1 for r in regs if r["tipo"] == "salida")
    return jsonify({"registros": regs, "total": len(regs), "entradas": ent, "salidas": sal, "fecha": today_local().strftime("%d/%m/%Y")})


# --- CHECK-IN ---
@app.route("/checkin")
def checkin():
    t = request.args.get("token", "")
    if not qr_validar(t):
        return render_template("confirmacion.html", exito=False, mensaje="El codigo QR ha expirado. Escanea el QR actual.")
    return render_template("checkin.html", token_qr=t)


@app.route("/api/checkin", methods=["POST"])
def api_checkin():
    data = request.get_json()
    if not data:
        return jsonify({"ok": False, "mensaje": "Datos invalidos."}), 400
    tqr = data.get("token_qr", "")
    tdev = data.get("token_dispositivo", "")
    if not qr_validar(tqr):
        return jsonify({"ok": False, "mensaje": "QR expirado."}), 400
    if not tdev:
        return jsonify({"ok": False, "mensaje": "Dispositivo no registrado."}), 400
    emp_id = device_validar(tdev)
    if not emp_id:
        return jsonify({"ok": False, "mensaje": "Token invalido."}), 400
    emp = db_obtener_empleado(emp_id)
    if not emp or not emp["activo"]:
        return jsonify({"ok": False, "mensaje": "Empleado no encontrado o inactivo."}), 400
    if emp["token_dispositivo"] != tdev:
        return jsonify({"ok": False, "mensaje": "Dispositivo no vinculado."}), 400

    # Una sola consulta con las marcas de hoy: sirve para el anti-rebote, para
    # saber si la jornada ya esta completa y para decidir el tipo.
    regs_hoy = db_registros_hoy_empleado(emp_id)

    # Recarga / doble escaneo: no se crea un registro nuevo, se repite el anterior.
    ventana = db_get_antirrebote()
    if regs_hoy and ventana > 0:
        previo = regs_hoy[-1]
        seg = (now_local() - to_local(previo["fecha_hora"])).total_seconds()
        if 0 <= seg < ventana:
            hora = to_local(previo["fecha_hora"]).strftime("%H:%M:%S")
            return jsonify({
                "ok": True, "duplicado": True, "nombre": emp["nombre"],
                "tipo": previo["tipo"], "hora": hora,
                "mensaje": f"Ya habias registrado tu {previo['tipo']} a las {hora}.",
            })

    # El almuerzo no se marca: dos marcas por dia y nada mas. Un tercer escaneo
    # antes creaba una entrada huerfana que dañaba el dia entero.
    entradas = [r for r in regs_hoy if r["tipo"] == "entrada"]
    salidas = [r for r in regs_hoy if r["tipo"] == "salida"]
    if entradas and salidas:
        primera = to_local(entradas[0]["fecha_hora"])
        ultima = to_local(salidas[-1]["fecha_hora"])
        duracion = (ultima - primera).total_seconds() / 60

        if duracion >= DUPLICADO_MAX_MINUTOS:
            return jsonify({
                "ok": False, "jornada_completa": True, "nombre": emp["nombre"],
                "mensaje": f"Tu jornada de hoy ya esta registrada: entrada {primera:%H:%M} y "
                           f"salida {ultima:%H:%M}. El almuerzo no se marca. "
                           f"Si algo esta mal, avisa a Recursos Humanos.",
            }), 409

        # Entrada y salida a minutos de distancia: no es una jornada, es un
        # doble escaneo. Se reubica esa salida a la hora real en vez de dejar
        # a la persona bloqueada con un dia de cero horas.
        ahora = now_local()
        db_mover_salida(salidas[-1]["id"], ahora)
        return jsonify({
            "ok": True, "duplicado": False, "corregido": True,
            "nombre": emp["nombre"], "tipo": "salida",
            "hora": ahora.strftime("%H:%M:%S"),
            "mensaje": "Salida registrada.",
            "recordatorio": f"Se corrigio un registro duplicado de las {ultima:%H:%M}.",
        })

    tipo = ("entrada" if regs_hoy[-1]["tipo"] == "salida" else "salida") if regs_hoy else tipo_por_hora(now_local())
    creado = db_registrar_asistencia(emp_id, tipo, tqr)
    hora = to_local(creado["fecha_hora"]).strftime("%H:%M:%S") if creado and creado.get("fecha_hora") else now_local().strftime("%H:%M:%S")
    return jsonify({
        "ok": True, "duplicado": False, "nombre": emp["nombre"],
        "tipo": tipo, "hora": hora,
        "mensaje": f"{tipo.capitalize()} registrada.",
        # El olvido de la salida es lo que mas ensucia los reportes: se avisa
        # en el momento, que es cuando la persona todavia puede hacer algo.
        "recordatorio": "No olvides marcar tu SALIDA al terminar la jornada." if tipo == "entrada" else "",
    })


# --- DEVICE REGISTRATION ---
@app.route("/registro-dispositivo")
def registro_dispositivo():
    t = request.args.get("token", "")
    emp_id = reg_validar(t)
    if not emp_id:
        return render_template("confirmacion.html", exito=False, mensaje="Enlace invalido o expirado.")
    emp = db_obtener_empleado(emp_id)
    if not emp:
        return render_template("confirmacion.html", exito=False, mensaje="Empleado no encontrado.")
    return render_template("registro_dispositivo.html", empleado=emp, token_reg=t)


@app.route("/api/registro-dispositivo", methods=["POST"])
def api_registro_dispositivo():
    data = request.get_json()
    if not data:
        return jsonify({"ok": False, "mensaje": "Datos invalidos."}), 400
    t = data.get("token_reg", "")
    emp_id = reg_validar(t)
    if not emp_id:
        return jsonify({"ok": False, "mensaje": "Token invalido."}), 400
    emp = db_obtener_empleado(emp_id)
    if not emp:
        return jsonify({"ok": False, "mensaje": "Empleado no encontrado."}), 400
    tok = device_token(emp_id)
    db_vincular(emp_id, tok)
    return jsonify({"ok": True, "mensaje": f"Vinculado para {emp['nombre']}.", "token_dispositivo": tok, "nombre": emp["nombre"]})


# --- ADMIN ---
@app.route("/admin/login", methods=["GET", "POST"])
def admin_login():
    if request.method == "POST":
        if db_verificar_password(request.form.get("password", "")):
            session["admin"] = True
            return redirect(url_for("admin_panel"))
        return render_template("admin_login.html", error="Contrasena incorrecta.")
    return render_template("admin_login.html")


@app.route("/admin")
@admin_required
def admin_panel():
    return render_template("admin.html")


@app.route("/admin/logout", methods=["POST"])
def admin_logout():
    session.pop("admin", None)
    return redirect(url_for("index"))


@app.route("/api/admin/empleados")
@admin_required
def api_admin_empleados():
    return jsonify({"empleados": db_listar_empleados(solo_activos=False)})


@app.route("/api/admin/empleados", methods=["POST"])
@admin_required
def api_admin_crear():
    data = request.get_json()
    if not data or not data.get("nombre", "").strip():
        return jsonify({"ok": False, "mensaje": "Nombre obligatorio."}), 400
    eid = db_crear_empleado(
        data["nombre"].strip(),
        data.get("departamento", "").strip(),
        data.get("hora_entrada", "09:00").strip(),
        data.get("hora_salida", "18:00").strip(),
    )
    return jsonify({"ok": True, "id": eid})


@app.route("/api/admin/empleados/<int:eid>", methods=["PUT"])
@admin_required
def api_admin_editar(eid):
    data = request.get_json()
    if not data:
        return jsonify({"ok": False, "mensaje": "Datos invalidos."}), 400
    db_actualizar_empleado(
        eid,
        nombre=data.get("nombre"),
        departamento=data.get("departamento"),
        hora_entrada=data.get("hora_entrada"),
        hora_salida=data.get("hora_salida"),
    )
    return jsonify({"ok": True})


@app.route("/api/admin/empleados/<int:eid>/toggle", methods=["POST"])
@admin_required
def api_admin_toggle(eid):
    emp = db_obtener_empleado(eid)
    if not emp:
        return jsonify({"ok": False, "mensaje": "No encontrado."}), 404
    new = 0 if emp["activo"] else 1
    db_actualizar_empleado(eid, activo=new)
    return jsonify({"ok": True, "activo": new})


@app.route("/api/admin/empleados/<int:eid>/qr-registro")
@admin_required
def api_admin_qr(eid):
    emp = db_obtener_empleado(eid)
    if not emp:
        return jsonify({"ok": False, "mensaje": "No encontrado."}), 404
    url = qr_registro_url(eid)
    return jsonify({"ok": True, "qr_base64": qr_base64(url), "nombre": emp["nombre"]})


@app.route("/api/admin/empleados/<int:eid>/desvincular", methods=["POST"])
@admin_required
def api_admin_desvincular(eid):
    emp = db_obtener_empleado(eid)
    if not emp:
        return jsonify({"ok": False, "mensaje": "No encontrado."}), 404
    db_desvincular(eid)
    return jsonify({"ok": True})


@app.route("/api/admin/config", methods=["GET"])
@admin_required
def api_admin_get_config():
    reglas = db_get_reglas_extras()
    return jsonify({
        "nombre_empresa": db_get_config("nombre_empresa") or "NEVOX FARMA",
        "tolerancia_minutos": db_get_config("tolerancia_minutos") or "15",
        "horario_semanal": db_get_horario_semanal(),
        "dias_semana": DIAS_SEMANA,
        "extras_minimo_minutos": reglas["minimo"],
        "extras_redondeo_minutos": reglas["redondeo"],
        "checkin_antirrebote_segundos": db_get_antirrebote(),
        "jornada_minima_minutos": db_get_jornada_minima(),
    })


@app.route("/api/admin/config", methods=["POST"])
@admin_required
def api_admin_save_config():
    data = request.get_json()
    if not data:
        return jsonify({"ok": False, "mensaje": "Datos invalidos."}), 400
    if "nombre_empresa" in data:
        db_set_config("nombre_empresa", data["nombre_empresa"])
    if "tolerancia_minutos" in data:
        try:
            t = int(data["tolerancia_minutos"])
            if t < 0:
                raise ValueError
            db_set_config("tolerancia_minutos", str(t))
        except ValueError:
            return jsonify({"ok": False, "mensaje": "Tolerancia invalida."}), 400
    for clave, etiqueta in [("extras_minimo_minutos", "minimo de hora extra"),
                            ("extras_redondeo_minutos", "redondeo de hora extra"),
                            ("checkin_antirrebote_segundos", "anti-rebote de check-in"),
                            ("jornada_minima_minutos", "jornada minima")]:
        if clave in data:
            try:
                v = int(data[clave])
                if v < 0:
                    raise ValueError
                db_set_config(clave, str(v))
            except (TypeError, ValueError):
                return jsonify({"ok": False, "mensaje": f"Valor invalido para {etiqueta}."}), 400
    if "horario_semanal" in data:
        horario = normalizar_horario(data["horario_semanal"])
        for i in range(7):
            turno = horario[str(i)]
            if turno and turno["salida"] <= turno["entrada"]:
                return jsonify({
                    "ok": False,
                    "mensaje": f"{DIAS_SEMANA[i]}: la hora de salida debe ser mayor que la de entrada.",
                }), 400
        db_set_horario_semanal(horario)
    if data.get("nuevo_password"):
        if data["nuevo_password"] != data.get("confirmar_password"):
            return jsonify({"ok": False, "mensaje": "No coinciden."}), 400
        if len(data["nuevo_password"]) < 4:
            return jsonify({"ok": False, "mensaje": "Min 4 caracteres."}), 400
        db_cambiar_password(data["nuevo_password"])
    return jsonify({"ok": True, "mensaje": "Configuracion guardada."})


# --- CORRECCION DE REGISTROS ---
TIPOS_VALIDOS = ("entrada", "salida")


def _valida_registro(data):
    """Devuelve (fecha, hora, tipo) o (None, mensaje de error)."""
    fecha = (data.get("fecha") or "").strip()
    hora = (data.get("hora") or "").strip()
    tipo = (data.get("tipo") or "").strip().lower()
    if tipo not in TIPOS_VALIDOS:
        return None, "El tipo debe ser entrada o salida."
    try:
        date.fromisoformat(fecha)
    except ValueError:
        return None, "Fecha invalida."
    if not _hhmm_valido(hora):
        return None, "Hora invalida."
    return (fecha, hora, tipo), None


@app.route("/api/admin/registros")
@admin_required
def api_admin_registros():
    desde = request.args.get("desde") or today_local().isoformat()
    hasta = request.args.get("hasta") or desde
    eid = request.args.get("empleado_id")
    regs = db_registros_rango(desde, hasta, int(eid) if eid else None)
    for r in regs:
        loc = datetime.fromisoformat(r["fecha_hora"])
        r["fecha_dia"] = loc.strftime("%Y-%m-%d")
        r["hora_corta"] = loc.strftime("%H:%M")
    return jsonify({"registros": regs, "desde": desde, "hasta": hasta})


@app.route("/api/admin/dias-por-corregir")
@admin_required
def api_admin_dias_por_corregir():
    desde, hasta = _rango_args()
    eid = request.args.get("empleado_id")
    return jsonify({
        "dias": db_dias_por_corregir(desde, hasta, int(eid) if eid else None),
        "desde": desde, "hasta": hasta,
    })


@app.route("/api/admin/registros", methods=["POST"])
@admin_required
def api_admin_crear_registro():
    data = request.get_json() or {}
    try:
        emp_id = int(data.get("empleado_id"))
    except (TypeError, ValueError):
        return jsonify({"ok": False, "mensaje": "Selecciona un empleado."}), 400
    if not db_obtener_empleado(emp_id):
        return jsonify({"ok": False, "mensaje": "Empleado no encontrado."}), 404
    valores, error = _valida_registro(data)
    if error:
        return jsonify({"ok": False, "mensaje": error}), 400
    db_crear_registro(emp_id, *valores)
    return jsonify({"ok": True, "mensaje": "Registro agregado."})


@app.route("/api/admin/registros/<int:reg_id>", methods=["PUT"])
@admin_required
def api_admin_editar_registro(reg_id):
    data = request.get_json() or {}
    valores, error = _valida_registro(data)
    if error:
        return jsonify({"ok": False, "mensaje": error}), 400
    fecha, hora, tipo = valores
    db_actualizar_registro(reg_id, fecha, hora, tipo)
    return jsonify({"ok": True, "mensaje": "Registro actualizado."})


@app.route("/api/admin/registros/<int:reg_id>", methods=["DELETE"])
@admin_required
def api_admin_eliminar_registro(reg_id):
    db_eliminar_registro(reg_id)
    return jsonify({"ok": True, "mensaje": "Registro eliminado."})


@app.route("/api/admin/limpiar-registros", methods=["POST"])
@admin_required
def api_admin_limpiar_reg():
    db_limpiar_registros()
    return jsonify({"ok": True, "mensaje": "Registros eliminados."})


@app.route("/api/admin/limpiar-todo", methods=["POST"])
@admin_required
def api_admin_limpiar_todo():
    db_limpiar_todo()
    return jsonify({"ok": True, "mensaje": "Registros y empleados eliminados."})


# --- REPORTS ---
@app.route("/reportes")
def reportes():
    return render_template("reports.html")


def _rango_args():
    return (
        request.args.get("desde", today_local().replace(day=1).isoformat()),
        request.args.get("hasta", today_local().isoformat()),
    )


@app.route("/api/reportes/areas")
def api_reportes_areas():
    return jsonify({"areas": db_listar_areas()})


@app.route("/api/reportes/horas")
def api_reportes_horas():
    desde, hasta = _rango_args()
    area = request.args.get("departamento") or None
    data = db_resumen_periodo(desde, hasta, departamento=area)
    return jsonify({
        "datos": data["resumen"], "reglas": data["reglas"],
        "desde": desde, "hasta": hasta,
    })


@app.route("/api/reportes/horas-extras")
def api_reportes_horas_extras():
    desde, hasta = _rango_args()
    eid = request.args.get("empleado_id")
    area = request.args.get("departamento") or None
    data = db_resumen_periodo(desde, hasta, int(eid) if eid else None, area)
    detalle = data["detalle"]
    if request.args.get("solo_extras", "1") == "1":
        detalle = [d for d in detalle if d["extra_min"] > 0 or d["revisar"]]
    return jsonify({
        "detalle": detalle, "resumen": data["resumen"],
        "horario": data["horario"], "reglas": data["reglas"],
        "desde": desde, "hasta": hasta,
    })


@app.route("/api/reportes/retardos")
def api_reportes_retardos():
    desde, hasta = _rango_args()
    area = request.args.get("departamento") or None
    datos, sin_corregir = db_retardos(desde, hasta, area)
    return jsonify({
        "datos": datos,
        "sin_corregir": sin_corregir,
        "tolerancia": int(db_get_config("tolerancia_minutos") or "15"),
        "desde": desde, "hasta": hasta,
    })


@app.route("/api/reportes/exportar-excel")
def api_exportar_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    desde, hasta = _rango_args()
    eid = request.args.get("empleado_id")
    if eid:
        eid = int(eid)
    area_filtro = request.args.get("departamento") or None
    regs = db_registros_rango(desde, hasta, eid)
    if area_filtro:
        regs = [r for r in regs if (r["departamento"] or SIN_AREA) == area_filtro]
    periodo = db_resumen_periodo(desde, hasta, eid, area_filtro)
    retardos, ret_sin_corregir = db_retardos(desde, hasta, area_filtro)
    if eid:
        retardos = [r for r in retardos if r["empleado_id"] == eid]

    hf = Font(bold=True, color="FFFFFF", size=11)
    hfill = PatternFill(start_color="ea8511", end_color="ea8511", fill_type="solid")
    tf = Font(bold=True, size=14, color="1d120e")
    c = Alignment(horizontal="center", vertical="center")
    b = Border(
        left=Side(style="thin", color="e0e0e2"),
        right=Side(style="thin", color="e0e0e2"),
        top=Side(style="thin", color="e0e0e2"),
        bottom=Side(style="thin", color="e0e0e2"),
    )
    af = PatternFill(start_color="f7f7f8", end_color="f7f7f8", fill_type="solid")
    # Estilos por tipo de fila: cabecera de area, subtotal y retardo en rojo.
    grupo_font = Font(bold=True, size=12, color="1d120e")
    grupo_fill = PatternFill(start_color="fef7ed", end_color="fef7ed", fill_type="solid")
    sub_font = Font(bold=True, color="1d120e")
    sub_fill = PatternFill(start_color="f0f0f2", end_color="f0f0f2", fill_type="solid")
    # ARGB explicito: es la marca que pidio el cliente, no depende del relleno
    # de canal alfa que haga openpyxl.
    rojo_font = Font(bold=True, color="FFDC2626")
    rojo_fill = PatternFill(start_color="FFFEF2F2", end_color="FFFEF2F2", fill_type="solid")

    def armar_hoja(ws, titulo, cabeceras, filas, anchos):
        """filas: lista de dicts {"tipo": grupo|dato|alerta|subtotal, "vals": [...]}"""
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(cabeceras))
        ws.cell(row=1, column=1, value=titulo).font = tf
        for col, h in enumerate(cabeceras, 1):
            cell = ws.cell(row=3, column=col, value=h)
            cell.font = hf
            cell.fill = hfill
            cell.alignment = c
            cell.border = b
        for i, fila in enumerate(filas, 4):
            tipo = fila["tipo"]
            for col in range(1, len(cabeceras) + 1):
                val = fila["vals"][col - 1] if col <= len(fila["vals"]) else None
                cell = ws.cell(row=i, column=col, value=val)
                cell.border = b
                if tipo == "grupo":
                    cell.font = grupo_font
                    cell.fill = grupo_fill
                elif tipo == "subtotal":
                    cell.font = sub_font
                    cell.fill = sub_fill
                elif tipo == "alerta":
                    cell.font = rojo_font
                    cell.fill = rojo_fill
                elif i % 2 == 0:
                    cell.fill = af
            # El merge va despues de escribir: openpyxl deja de solo lectura
            # las celdas fusionadas que no son la superior izquierda.
            if tipo == "grupo":
                ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=len(cabeceras))
        ws.freeze_panes = "A4"
        for col, ancho in enumerate(anchos, 1):
            ws.column_dimensions[ws.cell(row=3, column=col).column_letter].width = ancho

    def agrupar(items, vals_fn, subtotal_fn=None, alerta_fn=None):
        """Convierte items (ya ordenados por area) en filas con cabecera de
        area y, opcionalmente, una fila de subtotal por area."""
        filas, area_actual, grupo = [], None, []
        for it in items:
            area = it.get("departamento") or SIN_AREA
            if area != area_actual:
                if grupo and subtotal_fn:
                    filas.append({"tipo": "subtotal", "vals": subtotal_fn(area_actual, grupo)})
                area_actual, grupo = area, []
                filas.append({"tipo": "grupo", "vals": [f"AREA: {area}"]})
            grupo.append(it)
            filas.append({
                "tipo": "alerta" if (alerta_fn and alerta_fn(it)) else "dato",
                "vals": vals_fn(it),
            })
        if grupo and subtotal_fn:
            filas.append({"tipo": "subtotal", "vals": subtotal_fn(area_actual, grupo)})
        return filas

    wb = Workbook()

    # --- Hoja 1: registros crudos, agrupados por area ---
    regs_ord = sorted(regs, key=lambda r: ((r["departamento"] or SIN_AREA), r["nombre"], r["fecha_hora"]))
    armar_hoja(
        wb.active, f"NEVOX FARMA - Registros {desde} al {hasta}",
        ["Fecha", "Hora", "Empleado", "Area", "Tipo"],
        agrupar(regs_ord, lambda r: [
            datetime.fromisoformat(r["fecha_hora"]).strftime("%Y-%m-%d"),
            datetime.fromisoformat(r["fecha_hora"]).strftime("%H:%M:%S"),
            r["nombre"], r["departamento"] or SIN_AREA, r["tipo"].upper(),
        ]),
        [14, 12, 28, 22, 12],
    )
    wb.active.title = "Registros"

    # --- Hoja 2: detalle diario de horas extras ---
    def _nota(d):
        notas = []
        if d["no_laboral"]:
            notas.append("Dia no laboral")
        if d["revisar"]:
            notas.append(d["motivo"])
        return " / ".join(notas)

    detalle_ext = [d for d in periodo["detalle"] if d["extra_bruto_min"] > 0 or d["revisar"]]
    armar_hoja(
        wb.create_sheet("Horas Extras"),
        f"NEVOX FARMA - Horas extras {desde} al {hasta} "
        f"(minimo {periodo['reglas']['minimo']} min, redondeo {periodo['reglas']['redondeo']} min)",
        ["Empleado", "Area", "Fecha", "Dia", "Entrada", "Salida",
         "Salida programada", "Horas trabajadas", "Extra registrada (min)",
         "Extra validada (min)", "Extra validada (hrs)", "Observacion"],
        agrupar(
            detalle_ext,
            lambda d: [
                d["nombre"], d["departamento"], d["fecha"], d["dia"],
                d["entrada"], d["salida"], d["salida_programada"] or "-",
                d["horas_trabajadas"], d["extra_bruto_min"], d["extra_min"],
                d["extra_horas"], _nota(d),
            ],
            lambda area, g: [
                f"Subtotal {area}", "", "", "", "", "", "",
                round(sum(x["horas_trabajadas"] for x in g), 2), "",
                sum(x["extra_min"] for x in g),
                round(sum(x["extra_horas"] for x in g), 2), "",
            ],
        ),
        [26, 20, 12, 12, 10, 10, 18, 16, 20, 18, 18, 24],
    )

    # --- Hoja 3: retardos, con la fila en rojo cuando excede la tolerancia ---
    tol = int(db_get_config("tolerancia_minutos") or "15")
    retardos_ord = sorted(retardos, key=lambda r: (r["departamento"], r["nombre"], r["fecha"]))
    armar_hoja(
        wb.create_sheet("Retardos"),
        f"NEVOX FARMA - Retardos {desde} al {hasta} (tolerancia {tol} min; "
        f"en rojo los que la exceden)"
        + (f" - {ret_sin_corregir} dia(s) excluidos por marcas sin corregir"
           if ret_sin_corregir else ""),
        ["Empleado", "Area", "Fecha", "Dia", "Hora programada", "Hora registro",
         "Minutos tarde", "Estado"],
        agrupar(
            retardos_ord,
            lambda r: [
                r["nombre"], r["departamento"], r["fecha"], r["dia"],
                r["hora_programada"], r["hora_registro"], r["minutos_tarde"],
                "Dentro de tolerancia" if r["con_tolerancia"] else "RETARDO",
            ],
            lambda area, g: [
                f"Subtotal {area}", "", "", "", "",
                f"{sum(1 for x in g if not x['con_tolerancia'])} retardos",
                sum(x["minutos_tarde"] for x in g), f"{len(g)} llegadas tarde",
            ],
            alerta_fn=lambda r: not r["con_tolerancia"],
        ),
        [26, 20, 12, 12, 18, 16, 14, 22],
    )

    # --- Hoja 4: resumen por empleado ---
    armar_hoja(
        wb.create_sheet("Resumen"),
        f"NEVOX FARMA - Resumen {desde} al {hasta}",
        ["Empleado", "Area", "Horas trabajadas", "Horas extras",
         "Extras dias habiles", "Extras fin de semana", "Dias con extra",
         "Dias sin salida"],
        agrupar(
            periodo["resumen"],
            lambda r: [
                r["nombre"], r["departamento"], r["horas"], r["extras_horas"],
                r["extras_habil_horas"], r["extras_no_laboral_horas"],
                r["dias_con_extra"], r["dias_incompletos"],
            ],
            lambda area, g: [
                f"Subtotal {area}", "",
                round(sum(x["horas"] for x in g), 2),
                round(sum(x["extras_horas"] for x in g), 2),
                round(sum(x["extras_habil_horas"] for x in g), 2),
                round(sum(x["extras_no_laboral_horas"] for x in g), 2),
                sum(x["dias_con_extra"] for x in g),
                sum(x["dias_incompletos"] for x in g),
            ],
        ),
        [26, 20, 18, 14, 18, 20, 16, 16],
    )

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf, as_attachment=True,
        download_name=f"registros_{desde}_{hasta}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
